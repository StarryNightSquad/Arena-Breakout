import time
print("=" * 50)
print("本程序由繁星攻略组制作")
print("QQ：1035835099")
print("WX：TangTangMei18")
print("=" * 50)
print(" ")

time.sleep(5) # 延时2秒

import openpyxl
from openpyxl import load_workbook
import os
import re
from decimal import Decimal, getcontext, ROUND_HALF_UP

class ArmorPenetrationCalculator:
    def __init__(self):
        # 设置Decimal精度
        getcontext().prec = 10  # 设置足够高的精度
        
        # 加载所有Excel文件
        print("正在加载数据文件...")
        
        # 加载穿透数据
        self.penetration_data = {}
        self.load_penetration_data()
        
        # 加载武器数据
        self.weapons_by_category = {}
        self.weapons_data = {}
        self.load_weapon_data()
        
        # 加载弹药数据
        self.ammo_by_caliber = {}
        self.ammo_data = {}
        self.load_ammo_data()
        
        # 加载护甲数据
        self.armor_by_level = {}
        self.armor_data = {}
        self.load_armor_data()
        
        print("数据加载完成！\n")
    
    def standardize_caliber(self, caliber_str):
        """标准化口径字符串，使其能够正确匹配"""
        if not caliber_str:
            return ""
        
        # 转换为字符串并标准化格式
        caliber = str(caliber_str).strip()
        
        # 统一大小写
        caliber = caliber.upper()
        
        # 替换常见的变体
        caliber = caliber.replace('X', 'x')  # 统一使用小写x
        caliber = caliber.replace(' ', '')   # 移除空格
        caliber = caliber.replace('MM', 'mm')  # 统一mm为小写
        
        # 处理特定格式
        if caliber == '5.56X45MM':
            caliber = '5.56x45mm'
        elif caliber == '5.45X39MM':
            caliber = '5.45x39mm'
        elif caliber == '7.62X39MM':
            caliber = '7.62x39mm'
        elif caliber == '7.62X51MM':
            caliber = '7.62x51mm'
        elif caliber == '7.62X54MM':
            caliber = '7.62x54mm'
        elif caliber == '5.8X42MM':
            caliber = '5.8x42mm'
        elif caliber == '9X19MM':
            caliber = '9x19mm'
        elif caliber == '9X39MM':
            caliber = '9x39mm'
        elif caliber == '12 GAUGE':
            caliber = '12 Gauge'
        elif caliber == '.45 ACP':
            caliber = '.45 ACP'
        elif caliber == '.44 MAG':
            caliber = '.44 Mag'
        elif caliber == '.338 LAPUA MAG':
            caliber = '.338 Lapua Mag'
        
        return caliber
    
    def load_penetration_data(self):
        """加载穿透百分比数据"""
        try:
            wb = load_workbook('穿透数据整理结果.xlsx')
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    k_value = int(row[0])
                    # 使用Decimal存储穿透百分比
                    self.penetration_data[k_value] = Decimal(str(row[3]))
            
            print(f"已加载 {len(self.penetration_data)} 条穿透数据")
            
        except Exception as e:
            print(f"加载穿透数据失败: {e}")
            # 使用示例数据作为备选
            self.penetration_data = {
                -10: Decimal('0.6'), -9: Decimal('0.6053'), -8: Decimal('0.611'), 
                -7: Decimal('0.615'), -6: Decimal('0.62'), -5: Decimal('0.626'), 
                -4: Decimal('0.628'), -3: Decimal('0.639'), -1: Decimal('0.645'), 
                0: Decimal('0.65'), 1: Decimal('0.6556'), 2: Decimal('0.66'), 
                3: Decimal('0.666'), 4: Decimal('0.67'), 5: Decimal('0.675'),
                6: Decimal('0.681'), 7: Decimal('0.685'), 8: Decimal('0.688'), 
                9: Decimal('0.696'), 10: Decimal('0.9'), 11: Decimal('0.91'), 
                12: Decimal('0.92'), 13: Decimal('0.93'), 14: Decimal('0.94'), 
                15: Decimal('0.95'), 17: Decimal('0.97'), 18: Decimal('0.98')
            }
    
    def load_weapon_data(self):
        """加载武器数据"""
        try:
            wb = load_workbook('武器数据.xlsx')
            ws = wb.active
            
            current_category = None
            weapon_id = 0
            
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and not row[1]:  # 武器类别行（如"突击步枪"）
                    current_category = row[0]
                    self.weapons_by_category[current_category] = []
                elif row[0] and row[1]:  # 武器名称行
                    weapon_id += 1
                    
                    # 标准化口径
                    caliber = self.standardize_caliber(row[1])
                    
                    weapon = {
                        'id': weapon_id,
                        'name': row[0],
                        'category': current_category,
                        'caliber': caliber,
                        'original_caliber': row[1],  # 保存原始口径用于显示
                        'barrels': []
                    }
                    
                    # 添加第一个枪管
                    barrel_name = row[4] if row[4] else "/"
                    barrel = {
                        'name': barrel_name,
                        'firing_power': Decimal(str(row[5])) if row[5] is not None else Decimal('0')
                    }
                    weapon['barrels'].append(barrel)
                    
                    self.weapons_by_category[current_category].append(weapon)
                    self.weapons_data[weapon_id] = weapon
                    
                elif not row[0] and row[4]:  # 额外的枪管行
                    # 找到最后一个添加的武器
                    weapon = self.weapons_by_category[current_category][-1]
                    barrel = {
                        'name': row[4],
                        'firing_power': Decimal(str(row[5])) if row[5] is not None else Decimal('0')
                    }
                    weapon['barrels'].append(barrel)
            
            print(f"已加载 {len(self.weapons_data)} 种武器，分为 {len(self.weapons_by_category)} 个类别")
            
        except Exception as e:
            print(f"加载武器数据失败: {e}")
    
    def load_ammo_data(self):
        """加载弹药数据"""
        try:
            wb = load_workbook('子弹数据.xlsx')
            ws = wb.active
            
            ammo_id = 0
            current_caliber = None
            
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0]:  # 新口径
                    # 标准化口径
                    current_caliber = self.standardize_caliber(row[0])
                    if current_caliber not in self.ammo_by_caliber:
                        self.ammo_by_caliber[current_caliber] = []
                
                if row[1]:  # 弹药型号
                    ammo_id += 1
                    
                    # 解析穿透等级
                    penetration_level = 0
                    if row[4]:  # D列是穿透等级（=ROUNDDOWN(F3/10,0)）
                        # 尝试从公式结果中提取数字
                        try:
                            # 如果是公式字符串，尝试计算
                            if isinstance(row[4], str) and row[4].startswith('='):
                                # 从F列获取值并除以10取整
                                row_num = row[4].split('F')[1].split('/')[0]
                                f_col_value = ws[f'F{int(row_num)}'].value
                                if f_col_value:
                                    penetration_level = int(float(f_col_value) // 10)
                            else:
                                # 直接转换
                                penetration_level = int(float(row[4]))
                        except:
                            # 如果解析失败，使用默认值
                            pass
                    
                    # 解析子弹穿透值（F列）
                    bullet_penetration = Decimal('0')
                    if row[5]:
                        try:
                            bullet_penetration = Decimal(str(row[5]))
                        except:
                            bullet_penetration = Decimal('0')
                    
                    ammo = {
                        'id': ammo_id,
                        'caliber': current_caliber,
                        'original_caliber': row[0],  # 保存原始口径用于显示
                        'name': row[1],
                        'damage': self.parse_damage_value(row[2]),
                        'armor_damage': Decimal(str(row[3])) if row[3] else Decimal('0'),
                        'penetration_level': penetration_level,
                        'bullet_penetration': bullet_penetration  # 保存子弹穿透值
                    }
                    
                    # 如果穿透等级为0，根据子弹穿透值计算
                    if ammo['penetration_level'] == 0 and ammo['bullet_penetration'] > 0:
                        ammo['penetration_level'] = int(ammo['bullet_penetration'] // 10)
                    
                    self.ammo_by_caliber[current_caliber].append(ammo)
                    self.ammo_data[ammo_id] = ammo
            
            # 打印所有加载的口径，方便调试
            print(f"已加载 {len(self.ammo_data)} 种子弹，涵盖 {len(self.ammo_by_caliber)} 种口径")
            print("可用口径:", ", ".join(sorted(self.ammo_by_caliber.keys())))
            
        except Exception as e:
            print(f"加载弹药数据失败: {e}")
    
    def parse_damage_value(self, damage_str):
        """解析伤害值（处理像"86*2"这样的格式）"""
        if not damage_str:
            return Decimal('0')
        
        damage_str = str(damage_str)
        if '*' in damage_str:
            parts = damage_str.split('*')
            try:
                return Decimal(str(parts[0])) * Decimal(str(parts[1]))
            except:
                return Decimal('0')
        else:
            try:
                return Decimal(str(damage_str))
            except:
                return Decimal('0')
    
    def load_armor_data(self):
        """加载护甲数据"""
        try:
            wb = load_workbook('护甲信息.xlsx')
            ws = wb.active
            
            armor_id = 0
            
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and row[1]:  # 护甲名称和等级
                    armor_id += 1
                    
                    # 解析耐久值
                    durability = Decimal('0')
                    if row[3]:
                        try:
                            if isinstance(row[3], (int, float)):
                                durability = Decimal(str(row[3]))
                            else:
                                # 如果是公式，尝试计算
                                if isinstance(row[3], str) and row[3].startswith('='):
                                    # 简单处理公式 =D3/E3
                                    try:
                                        d_col = row[3].split('=')[1].split('/')[0]
                                        e_col = row[3].split('=')[1].split('/')[1]
                                        # 这里简化处理，实际应该解析单元格引用
                                        # 由于数据中已经有计算好的值，我们可以使用后面的等效耐久列
                                        if row[5] and isinstance(row[5], (int, float)):
                                            durability = Decimal(str(row[5]))
                                    except:
                                        pass
                                else:
                                    durability = Decimal(str(row[3]).replace(',', '.'))
                        except:
                            durability = Decimal('0')
                    
                    # 解析损毁系数
                    destruction_coef = Decimal('0')
                    if row[4]:
                        try:
                            if isinstance(row[4], (int, float)):
                                destruction_coef = Decimal(str(row[4]))
                            else:
                                destruction_coef = Decimal(str(row[4]).replace(',', '.'))
                        except:
                            destruction_coef = Decimal('0')
                    
                    armor = {
                        'id': armor_id,
                        'name': row[0],
                        'level': int(row[1]),
                        'durability': durability,
                        'destruction_coef': destruction_coef,
                        'current_durability': durability  # 初始时当前耐久等于总耐久
                    }
                    
                    level = armor['level']
                    if level not in self.armor_by_level:
                        self.armor_by_level[level] = []
                    
                    self.armor_by_level[level].append(armor)
                    self.armor_data[armor_id] = armor
            
            print(f"已加载 {len(self.armor_data)} 件护甲")
            
        except Exception as e:
            print(f"加载护甲数据失败: {e}")
    
    def check_penetration_status(self, ammo_penetration_level, armor_level, armor_current_durability, armor_max_durability):
        """
        检查击穿状态
        返回: (can_penetrate, penetration_type)
        can_penetrate: True/False 是否能击穿
        penetration_type: "必定击穿护甲", "概率击穿护甲", "无法击穿护甲"
        """
        # 规则1: 当子弹穿透等级≥护甲等级，则必定击穿护甲
        if ammo_penetration_level >= armor_level:
            return True, "必定击穿护甲"
        
        # 规则2: 当子弹穿透等级＜护甲等级，则根据护甲耐久计算
        level_difference = armor_level - ammo_penetration_level
        
        # 计算耐久比例
        if armor_max_durability <= Decimal('0'):
            return True, "必定击穿护甲"  # 如果护甲最大耐久为0，直接击穿
        
        durability_ratio = armor_current_durability / armor_max_durability
        
        # 根据耐久比例判断击穿状态
        # 1 > 耐久比例 > 1/2: 概率被低一级子弹击穿
        # 1/2 ≥ 耐久比例 > 1/4: 必定被低一级子弹击穿，概率被低两级子弹击穿
        # 以此类推
        
        # 计算需要的耐久减半次数
        required_half_lives = level_difference
        
        # 根据耐久比例计算已经减半的次数
        import math
        if durability_ratio <= Decimal('0'):
            # 耐久为0或负数，视为已经减半无限次
            return True, "必定击穿护甲"
        
        # 使用Decimal计算log2，先将Decimal转换为float进行计算，再转换回Decimal
        # 注意：这里使用float可能会损失一些精度，但对于比较大小足够
        ratio_float = float(durability_ratio)
        if ratio_float > 0:
            half_lives_completed = math.floor(-math.log2(ratio_float))
        else:
            half_lives_completed = float('inf')
        
        if half_lives_completed >= required_half_lives:
            # 已经减半足够次数，必定击穿
            return True, "必定击穿护甲"
        elif half_lives_completed == required_half_lives - 1:
            # 只差一次减半，概率击穿
            return True, "概率击穿护甲"
        else:
            # 减半次数不足，无法击穿
            return False, "无法击穿护甲"
    
    def calculate_penetration_percentage(self, k_value):
        """根据K值计算穿透百分比"""
        k_value_int = int(k_value)
        
        if k_value_int < -10:
            return Decimal('0.6')  # K值小于-10时固定60%
        elif k_value_int > 20:
            return Decimal('1.0')  # K值大于20时固定100%
        
        # 找到最接近的K值
        if k_value_int in self.penetration_data:
            return self.penetration_data[k_value_int]
        else:
            # 如果K值不在表中，找最接近的值
            closest_k = min(self.penetration_data.keys(), key=lambda x: abs(x - k_value_int))
            return self.penetration_data[closest_k]
    
    def calculate_damage(self, weapon, barrel, ammo, armor, current_durability, current_max_durability=None):
        """计算伤害"""
        if current_max_durability is None:
            current_max_durability = armor['durability']
            
        # 检查击穿状态
        can_penetrate, penetration_type = self.check_penetration_status(
            ammo['penetration_level'],
            armor['level'],
            current_durability,
            current_max_durability
        )
        
        # 计算护甲损伤（精确到小数点后两位）
        armor_damage = (ammo['armor_damage'] * armor['destruction_coef']).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        new_durability = (current_durability - armor_damage).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # 确保剩余耐久不为负数
        if new_durability < Decimal('0'):
            new_durability = Decimal('0')
        
        if not can_penetrate:
            return {
                'can_penetrate': False,
                'penetration_type': penetration_type,
                'damage': Decimal('0'),
                'armor_damage_dealt': armor_damage,
                'remaining_durability': new_durability,
                'is_destroyed': new_durability <= Decimal('0'),
                'current_max_durability': current_max_durability  # 保存当前上限
            }
        
        # 如果可以击穿，计算详细伤害
        # 计算K值（使用子弹穿透值 - 护甲等级×10）
        k_value = ammo['bullet_penetration'] - Decimal(str(armor['level'])) * Decimal('10')
        
        # 获取穿透百分比
        penetration_percentage = self.calculate_penetration_percentage(k_value)
        
        # 计算伤害（精确到小数点后四位）
        base_damage = ammo['damage'] + barrel['firing_power']
        damage = (base_damage * penetration_percentage).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
        
        return {
            'can_penetrate': True,
            'penetration_type': penetration_type,
            'k_value': k_value,
            'penetration_percentage': penetration_percentage,
            'damage': damage,
            'armor_damage_dealt': armor_damage,
            'remaining_durability': new_durability,
            'is_destroyed': new_durability <= Decimal('0'),
            'current_max_durability': current_max_durability  # 保存当前上限
        }
    
    def select_weapon(self):
        """选择武器"""
        print("=" * 50)
        print("请选择武器类别:")
        
        categories = list(self.weapons_by_category.keys())
        for i, category in enumerate(categories, 1):
            weapon_count = len(self.weapons_by_category[category])
            print(f"{i}. {category} ({weapon_count}种武器)")
        
        while True:
            try:
                choice = int(input(f"\n请选择类别 (1-{len(categories)}): "))
                if 1 <= choice <= len(categories):
                    selected_category = categories[choice-1]
                    break
                else:
                    print("输入无效，请重新选择。")
            except ValueError:
                print("请输入数字。")
        
        print(f"\n{selected_category} 包含以下武器:")
        weapons = self.weapons_by_category[selected_category]
        
        for i, weapon in enumerate(weapons, 1):
            barrels_info = f" ({len(weapon['barrels'])}种枪管)" if len(weapon['barrels']) > 1 else ""
            # 显示原始口径
            display_caliber = weapon.get('original_caliber', weapon['caliber'])
            print(f"{i}. {weapon['name']} - {display_caliber}{barrels_info}")
        
        while True:
            try:
                weapon_choice = int(input(f"\n请选择武器 (1-{len(weapons)}): "))
                if 1 <= weapon_choice <= len(weapons):
                    selected_weapon = weapons[weapon_choice-1]
                    break
                else:
                    print("输入无效，请重新选择。")
            except ValueError:
                print("请输入数字。")
        
        return selected_weapon
    
    def select_barrel(self, weapon):
        """选择枪管"""
        print(f"\n{weapon['name']} 可用的枪管:")
        
        barrels = weapon['barrels']
        if len(barrels) == 1 and barrels[0]['name'] == "/":
            print("该武器只有默认枪管")
            return barrels[0]
        
        for i, barrel in enumerate(barrels, 1):
            power = barrel['firing_power']
            power_sign = "+" if power >= Decimal('0') else ""
            print(f"{i}. {barrel['name']} (击发威力: {power_sign}{power:.0f})")
        
        while True:
            try:
                choice = int(input(f"\n请选择枪管 (1-{len(barrels)}): "))
                if 1 <= choice <= len(barrels):
                    return barrels[choice-1]
                else:
                    print("输入无效，请重新选择。")
            except ValueError:
                print("请输入数字。")
    
    def select_ammo(self, caliber):
        """选择弹药"""
        # 标准化口径
        standardized_caliber = self.standardize_caliber(caliber)
        
        print(f"\n{caliber} 口径可用的弹药:")
        
        if standardized_caliber not in self.ammo_by_caliber:
            # 尝试查找所有可能匹配的口径
            print(f"警告: 未找到标准化口径 '{standardized_caliber}' 的弹药数据！")
            print(f"可用口径: {', '.join(self.ammo_by_caliber.keys())}")
            
            # 尝试查找相似口径
            matching_calibers = []
            for available_caliber in self.ammo_by_caliber.keys():
                if standardized_caliber in available_caliber or available_caliber in standardized_caliber:
                    matching_calibers.append(available_caliber)
            
            if matching_calibers:
                print(f"发现相似口径: {', '.join(matching_calibers)}")
                # 使用第一个匹配的口径
                standardized_caliber = matching_calibers[0]
                print(f"自动选择: {standardized_caliber}")
            else:
                print("无法找到匹配的口径弹药！")
                return None
        
        ammo_list = self.ammo_by_caliber[standardized_caliber]
        
        # 使用从1开始的连续编号显示弹药列表
        for i, ammo in enumerate(ammo_list, 1):
            # 显示原始口径
            display_caliber = ammo.get('original_caliber', ammo['caliber'])
            print(f"{i}. {ammo['name']} - 伤害: {ammo['damage']:.0f}, 穿甲等级: {ammo['penetration_level']}, 护甲伤害: {ammo['armor_damage']:.1f}")
        
        while True:
            try:
                choice = int(input(f"\n请选择弹药 (1-{len(ammo_list)}): "))
                if 1 <= choice <= len(ammo_list):
                    selected_ammo = ammo_list[choice-1]
                    return selected_ammo
                else:
                    print("输入无效，请重新选择。")
            except ValueError:
                print("请输入数字。")
    
    def select_armor(self):
        """选择护甲"""
        print("\n请选择护甲等级:")
        
        levels = sorted(self.armor_by_level.keys())
        for level in levels:
            armor_count = len(self.armor_by_level[level])
            print(f"{level}. 防护等级{level} ({armor_count}件)")
        
        while True:
            try:
                level = int(input(f"\n请选择防护等级 ({levels[0]}-{levels[-1]}): "))
                if level in levels:
                    break
                else:
                    print("输入无效，请重新选择。")
            except ValueError:
                print("请输入数字。")
        
        print(f"\n防护等级 {level} 的护甲:")
        armors = self.armor_by_level[level]
        
        for i, armor in enumerate(armors, 1):
            print(f"{i}. {armor['name']} - 耐久: {armor['durability']:.1f}")
        
        while True:
            try:
                choice = int(input(f"\n请选择护甲 (1-{len(armors)}): "))
                if 1 <= choice <= len(armors):
                    selected_armor = armors[choice-1]
                    break
                else:
                    print("输入无效，请重新选择。")
            except ValueError:
                print("请输入数字。")
        
        # 设置护甲当前上限和当前耐久
        print("\n" + "=" * 50)
        print("护甲耐久设置")
        print("=" * 50)
        print(f"护甲: {selected_armor['name']}")
        print(f"初始耐久上限: {selected_armor['durability']:.1f}")
        
        # 询问是否设置自定义耐久
        while True:
            custom_input = input(f"\n是否设置自定义耐久？(Y/N, 输入N则使用初始耐久上限和耐久): ").strip().upper()
            if custom_input == 'Y':
                # 设置当前上限
                while True:
                    try:
                        max_input = input(f"\n请输入当前护甲上限 (最多一位小数, 0-{selected_armor['durability']:.1f}): ").strip()
                        
                        # 检查输入格式
                        if not max_input.replace('.', '').isdigit():
                            print("输入必须为数字，请重新输入。")
                            continue
                        
                        # 检查小数位数
                        if '.' in max_input:
                            decimal_parts = max_input.split('.')
                            if len(decimal_parts) > 2:
                                print("输入格式错误，请重新输入。")
                                continue
                            if len(decimal_parts[1]) > 1:
                                print("最多允许一位小数，请重新输入。")
                                continue
                        
                        max_durability = Decimal(max_input)
                        
                        if Decimal('0') <= max_durability <= selected_armor['durability']:
                            current_max_durability = max_durability
                            break
                        else:
                            print(f"当前上限必须在0-{selected_armor['durability']:.1f}之间，请重新输入。")
                    except Exception as e:
                        print(f"输入无效: {e}，请重新输入。")
                
                # 设置当前耐久
                while True:
                    try:
                        durability_input = input(f"\n请输入当前护甲耐久 (最多一位小数, 0-{current_max_durability:.1f}): ").strip()
                        
                        # 检查输入格式
                        if not durability_input.replace('.', '').isdigit():
                            print("输入必须为数字，请重新输入。")
                            continue
                        
                        # 检查小数位数
                        if '.' in durability_input:
                            decimal_parts = durability_input.split('.')
                            if len(decimal_parts) > 2:
                                print("输入格式错误，请重新输入。")
                                continue
                            if len(decimal_parts[1]) > 1:
                                print("最多允许一位小数，请重新输入。")
                                continue
                        
                        current_durability = Decimal(durability_input)
                        
                        if Decimal('0') <= current_durability <= current_max_durability:
                            break
                        else:
                            print(f"当前耐久必须在0-{current_max_durability:.1f}之间，请重新输入。")
                    except Exception as e:
                        print(f"输入无效: {e}，请重新输入。")
                
                break
            elif custom_input == 'N':
                # 使用初始值
                current_max_durability = selected_armor['durability']
                current_durability = selected_armor['durability']
                print(f"使用初始耐久上限: {current_max_durability:.1f}")
                print(f"使用初始耐久: {current_durability:.1f}")
                break
            else:
                print("输入无效，请输入 Y 或 N。")
        
        # 将设置的值添加到护甲数据中
        selected_armor['current_max_durability'] = current_max_durability
        selected_armor['current_durability'] = current_durability
        
        # 显示设置完成信息
        print(f"\n设置完成:")
        print(f"当前护甲上限: {current_max_durability:.1f}")
        print(f"当前护甲耐久: {current_durability:.1f}")
        if current_max_durability > Decimal('0'):
            durability_ratio = (current_durability / current_max_durability).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
            print(f"耐久比例: {durability_ratio:.2%}")
        
        return selected_armor
    
    def simulate_shooting(self):
        """模拟射击直到护甲被击毁"""
        print("=" * 50)
        print(" ")
        print("=" * 50)
        
        # 选择武器、枪管、弹药、护甲
        weapon = self.select_weapon()
        barrel = self.select_barrel(weapon)
        
        # 显示武器原始口径用于选择弹药
        display_caliber = weapon.get('original_caliber', weapon['caliber'])
        ammo = self.select_ammo(display_caliber)
        
        if not ammo:
            print("无法继续计算，请重新开始。")
            return True  # 返回True表示可以开始新一轮计算
        
        armor = self.select_armor()
        
        # 获取用户设置的当前上限和当前耐久，如果没有设置则使用初始值
        current_max_durability = armor.get('current_max_durability', armor['durability'])
        current_durability = armor.get('current_durability', armor['durability'])
        
        # 显示初始信息
        print("\n" + "=" * 50)
        print("模拟开始")
        print("=" * 50)
        # 显示武器原始口径
        weapon_display_caliber = weapon.get('original_caliber', weapon['caliber'])
        print(f"武器: {weapon['name']} ({weapon_display_caliber})")
        print(f"枪管: {barrel['name']}")
        # 显示弹药原始口径
        ammo_display_caliber = ammo.get('original_caliber', ammo['caliber'])
        print(f"弹药: {ammo['name']} ({ammo_display_caliber}, 穿透等级: {ammo['penetration_level']})")
        print(f"护甲: {armor['name']} (防护等级: {armor['level']})")
        print(f"初始耐久上限: {armor['durability']:.1f}")
        print(f"当前耐久上限: {current_max_durability:.1f}")
        print(f"当前耐久: {current_durability:.1f}")
        if current_max_durability > Decimal('0'):
            durability_ratio = (current_durability / current_max_durability).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
            print(f"当前耐久比例: {durability_ratio:.2%}")
        print("=" * 50)
        
        # 模拟射击
        round_num = 1
        initial_max_durability = current_max_durability
        
        # 记录击穿状态
        first_probable_hit = None  # 第一次概率击穿的轮次
        first_guaranteed_hit = None  # 第一次必定击穿的轮次
        simulation_history = []  # 记录详细计算过程
        
        # 先进行完整模拟，记录所有结果
        while current_durability > Decimal('0'):
            # 计算本轮伤害
            result = self.calculate_damage(weapon, barrel, ammo, armor, current_durability, current_max_durability)
            
            # 记录结果
            simulation_history.append({
                'round': round_num,
                'penetration_type': result['penetration_type'],
                'damage': result['damage'],
                'remaining_durability': result['remaining_durability'],
                'armor_damage_dealt': result['armor_damage_dealt'],
                'can_penetrate': result['can_penetrate'],
                'current_durability_before': current_durability,
                'current_max_durability': current_max_durability
            })
            
            # 检查是否是第一次概率击穿
            if first_probable_hit is None and result['penetration_type'] == "概率击穿护甲":
                first_probable_hit = round_num
            
            # 检查是否是第一次必定击穿
            if first_guaranteed_hit is None and result['penetration_type'] == "必定击穿护甲":
                first_guaranteed_hit = round_num
            
            # 更新当前耐久
            current_durability = result['remaining_durability']
            
            # 检查护甲是否被击毁
            if result['is_destroyed']:
                break
            
            round_num += 1
        
        # 输出关键指标
        print("\n" + "=" * 50)
        print("击穿分析结果")
        print("=" * 50)
        
        if first_probable_hit is not None:
            print(f"概率击穿所需命中数：{first_probable_hit}")
        else:
            print("概率击穿所需命中数：未出现概率击穿")
        
        if first_guaranteed_hit is not None:
            print(f"稳定击穿所需命中数：{first_guaranteed_hit}")
        else:
            print("稳定击穿所需命中数：未出现稳定击穿")
        
        print(f"护甲在第{round_num}轮被彻底击毁！")
        
        # 询问是否显示详细计算过程
        print("\n" + "=" * 50)
        while True:
            show_details = input("是否显示详细计算过程？(Y/N): ").strip().upper()
            if show_details == 'Y':
                print("\n" + "=" * 50)
                print("详细计算过程")
                print("=" * 50)
                
                for i, result in enumerate(simulation_history):
                    if i < len(simulation_history) - 1:
                        print(f"\n第{result['round']}轮射击")
                        print(result['penetration_type'])
                        
                        # 只有击穿时才输出伤害
                        if result['can_penetrate'] and result['damage'] > Decimal('0'):
                            print(f"造成{result['damage']:.2f}伤害")  # 伤害输出保留两位小数
                        
                        # 输出剩余护甲耐久（输出保留一位小数）
                        print(f"剩余护甲耐久：{result['remaining_durability']:.1f}/{result['current_max_durability']:.1f}")
                        
                        # 显示耐久比例
                        if result['current_max_durability'] > Decimal('0'):
                            durability_ratio = (result['remaining_durability'] / result['current_max_durability']).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
                            print(f"耐久比例：{durability_ratio:.2%}")
                    else:
                        # 最后一轮，护甲被击毁
                        print(f"\n第{result['round']}轮射击")
                        print(result['penetration_type'])
                        
                        if result['can_penetrate'] and result['damage'] > Decimal('0'):
                            print(f"造成{result['damage']:.2f}伤害")
                        
                        print(f"剩余护甲耐久：{result['remaining_durability']:.1f}/{result['current_max_durability']:.1f}")
                        
                        # 显示耐久比例
                        if result['current_max_durability'] > Decimal('0'):
                            durability_ratio = (result['remaining_durability'] / result['current_max_durability']).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
                            print(f"耐久比例：{durability_ratio:.2%}")
                        
                        print("护甲被彻底击毁！")
                
                break
            elif show_details == 'N':
                break
            else:
                print("输入无效，请输入 Y 或 N。")
        
        # 恢复护甲原始耐久
        armor['current_durability'] = armor['durability']
        
        # 询问是否开始新一轮计算
        print("\n" + "=" * 50)
        while True:
            answer = input("是否开始新一轮计算？(Y/N): ").strip().upper()
            if answer == 'Y':
                print("\n" + "=" * 50)
                print("开始新一轮计算")
                print("=" * 50)
                return True  # 表示继续新一轮计算
            elif answer == 'N':
                return False  # 表示退出程序
            else:
                print("输入无效，请输入 Y 或 N。")
    
    def run(self):
        """运行主程序"""
        print("=" * 50)
        print("护甲系统模拟器")
        print("=" * 50)
        print("击穿规则:")
        print("1. 子弹穿透等级≥护甲等级：必定击穿护甲")
        print("2. 子弹穿透等级＜护甲等级：根据护甲耐久计算")
        print("3. 护甲耐久减半时，必定能被更低一级的子弹击穿")
        print("=" * 50)
        print("耐久比例计算：当前耐久/当前上限")
        print("如未设置自定义耐久，则使用初始耐久上限")
        print("=" * 50)
        
        # 直接开始模拟射击
        continue_simulation = True
        while continue_simulation:
            continue_simulation = self.simulate_shooting()
            # 如果simulate_shooting返回False，表示用户选择N，退出程序

def main():
    # 检查文件是否存在
    required_files = ['穿透数据整理结果.xlsx', '武器数据.xlsx', '子弹数据.xlsx', '护甲信息.xlsx']
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print("警告: 以下文件不存在:")
        for f in missing_files:
            print(f"  - {f}")
        print("请确保所有数据文件在当前目录下。")
        return
    
    # 创建并运行计算器
    calculator = ArmorPenetrationCalculator()
    calculator.run()

if __name__ == "__main__":
    main()
